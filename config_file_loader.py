import openpyxl
from typing import List
from agent import Agent

class ConfigFileLoader:
    def __init__(self,
                 config_file: str):
        self._wb = openpyxl.load_workbook(config_file)
        self.agents: List[Agent] = []
        self.map: List[List[str]] = []
        self._load_agent()
        self._load_map() 
    
    def _load_agent(self):
        agents_ws = self._wb['agents']
        for row in range(2, agents_ws.max_row + 1):
            for col in range(1, agents_ws.max_column + 1):
                id = str(agents_ws.cell(row, col).value)
                if col == 1:
                    agent = Agent(id)
                    self.agents.append(agent)

    def _load_map(self):
        map_ws = self._wb['map']
        for row in range(1, map_ws.max_row + 1):
            map_row = []
            for col in range(1, map_ws.max_column + 1):
                map_row.append(map_ws.cell(row, col).value)
            self.map.append(map_row)


if __name__ == "__main__":
    map_file = "./config/config.xlsx"
    config = ConfigFileLoader(map_file)
    print(config.agents)
